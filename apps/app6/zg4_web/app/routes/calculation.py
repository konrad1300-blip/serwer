from flask import Blueprint, render_template, request, flash, session, redirect, url_for, send_file
from app import db
from app.models import Grupa, Metoda, Obliczenie
from app.forms import ObliczeniaForm, WalidacjaForm
import json
import re
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

bp = Blueprint('calculation', __name__)

PRZEDZIALY = ["do 2m2", "od 2 do 20m2", "od 20 do 60m2", "powyżej 60m2"]

def waliduj_kod(kod):
    return re.match(r'^\d{3}-\d{4}-\d{3}$', kod) is not None

@bp.route('/', methods=['GET', 'POST'])
def index():
    form = ObliczeniaForm()
    grupy = Grupa.query.order_by(Grupa.nazwa).all()
    form.grupa_id.choices = [(g.id, g.nazwa) for g in grupy]

    if form.validate_on_submit():
        kod = form.kod.data
        if not waliduj_kod(kod):
            flash('Nieprawidłowy format kodu produktu.', 'danger')
            return render_template('calculation/index.html', form=form)

        grupa_id = form.grupa_id.data
        przedzial = form.przedzial.data
        grupa = Grupa.query.get(grupa_id)
        if not grupa:
            flash('Wybrana grupa nie istnieje.', 'danger')
            return render_template('calculation/index.html', form=form)

        session['obliczenia'] = {
            'kod': kod,
            'grupa_id': grupa_id,
            'przedzial': przedzial
        }
        return redirect(url_for('calculation.metry'))

    return render_template('calculation/index.html', form=form)

@bp.route('/laduj/<int:obliczenie_id>', methods=['GET'])
def laduj_do_obliczen(obliczenie_id):
    """Ładuje zapisane obliczenie do formularza obliczeń"""
    obliczenie = Obliczenie.query.get_or_404(obliczenie_id)
    
    # Zapisz dane w sesji
    session['obliczenia'] = {
        'kod': obliczenie.kod,
        'grupa_id': obliczenie.grupa_id,
        'przedzial': obliczenie.przedzial
    }
    
    # Zapisz wyniki w sesji
    wyniki = obliczenie.wyniki()
    session['wyniki'] = wyniki
    session['czas_calkowity'] = obliczenie.czas_calkowity
    session['kod'] = obliczenie.kod
    session['grupa_nazwa'] = obliczenie.grupa.nazwa
    session['przedzial'] = obliczenie.przedzial
    
    flash(f'Załadowano obliczenie z dnia {obliczenie.data.strftime("%Y-%m-%d %H:%M")}', 'success')
    return redirect(url_for('calculation.wynik'))

@bp.route('/metry', methods=['GET', 'POST'])
def metry():
    dane = session.get('obliczenia')
    if not dane:
        return redirect(url_for('calculation.index'))

    grupa = Grupa.query.get(dane['grupa_id'])
    if not grupa:
        flash('Grupa nie istnieje.', 'danger')
        return redirect(url_for('calculation.index'))

    if request.method == 'POST':
        metry_dane = {}
        wymuszenia = {}
        for metoda in grupa.metody:
            # Obsługa pustego pola lub błędnej wartości
            metry_str = request.form.get(f'metry_{metoda.id}')
            try:
                metry = float(metry_str) if metry_str else 0.0
            except ValueError:
                metry = 0.0

            if metry > 0:
                metry_dane[metoda.id] = metry
                if request.form.get(f'wymus_{metoda.id}') == 'on':
                    prac_str = request.form.get(f'prac_{metoda.id}')
                    try:
                        prac = int(prac_str) if prac_str else 1
                    except ValueError:
                        prac = 1
                    wymuszenia[metoda.id] = prac

        if not metry_dane:
            flash('Wprowadź przynajmniej jeden metraż.', 'danger')
        else:
            wyniki = []
            czas_calkowity = 0.0
            for metoda in grupa.metody:
                if metoda.id in metry_dane:
                    metry = metry_dane[metoda.id]
                    pracownicy, czas_na_metr = metoda.pobierz_czas(dane['przedzial'])
                    if metoda.id in wymuszenia:
                        pracownicy = wymuszenia[metoda.id]
                    czas_metody = metry * czas_na_metr * pracownicy
                    czas_calkowity += czas_metody
                    wyniki.append({
                        'nazwa': metoda.nazwa,
                        'metry': metry,
                        'czas_na_metr': czas_na_metr,
                        'pracownicy': pracownicy,
                        'czas_calkowity': czas_metody,
                        'czy_wymuszeni': metoda.id in wymuszenia
                    })
            session['wyniki'] = wyniki
            session['czas_calkowity'] = czas_calkowity
            session['kod'] = dane['kod']
            session['grupa_nazwa'] = grupa.nazwa
            session['przedzial'] = dane['przedzial']
            return redirect(url_for('calculation.wynik'))

    return render_template('calculation/metry.html', grupa=grupa, przedzial=dane['przedzial'])

@bp.route('/wynik', methods=['GET', 'POST'])
def wynik():
    wyniki = session.get('wyniki')
    czas_calkowity = session.get('czas_calkowity')
    kod = session.get('kod')
    grupa_nazwa = session.get('grupa_nazwa')
    przedzial = session.get('przedzial')

    if not wyniki:
        return redirect(url_for('calculation.index'))

    form = WalidacjaForm()
    if form.validate_on_submit():
        czas_produkcji = form.czas_produkcji.data
        odchylenie = ((czas_produkcji - czas_calkowity) / czas_calkowity) * 100 if czas_calkowity else 0
        status = ""
        if abs(odchylenie) <= 10:
            status = "W normie (≤10%)"
        elif abs(odchylenie) <= 20:
            status = "Dopuszczalne (≤20%)"
        else:
            status = "Poza normą (>20%)"
        flash(f'Odchylenie: {odchylenie:+.2f}% – {status}', 'info')
        return render_template('calculation/wynik.html', wyniki=wyniki, czas_calkowity=czas_calkowity,
                               kod=kod, grupa_nazwa=grupa_nazwa, przedzial=przedzial, form=form,
                               pokaz_walidacje=True, odchylenie=odchylenie, status=status)

    return render_template('calculation/wynik.html', wyniki=wyniki, czas_calkowity=czas_calkowity,
                           kod=kod, grupa_nazwa=grupa_nazwa, przedzial=przedzial, form=form,
                           pokaz_walidacje=False)

@bp.route('/zapisz', methods=['POST'])
def zapisz():
    wyniki = session.get('wyniki')
    czas_calkowity = session.get('czas_calkowity')
    kod = session.get('kod')
    grupa_nazwa = session.get('grupa_nazwa')
    przedzial = session.get('przedzial')

    if not wyniki or not kod:
        flash('Brak danych do zapisania.', 'danger')
        return redirect(url_for('calculation.index'))

    grupa = Grupa.query.filter_by(nazwa=grupa_nazwa).first()
    if not grupa:
        flash('Grupa nie istnieje.', 'danger')
        return redirect(url_for('calculation.wynik'))

    obliczenie = Obliczenie(
        kod=kod,
        grupa_id=grupa.id,
        przedzial=przedzial,
        czas_calkowity=czas_calkowity
    )
    obliczenie.zapisz_wyniki(wyniki)
    db.session.add(obliczenie)
    db.session.commit()

    flash('Obliczenie zostało zapisane.', 'success')
    return redirect(url_for('calculation.historia'))

@bp.route('/historia')
def historia():
    # Pobieranie parametrów filtrowania
    kod_filter = request.args.get('kod', '').strip()
    data_od = request.args.get('data_od', '').strip()
    data_do = request.args.get('data_do', '').strip()
    
    # Budowanie zapytania z filtrami
    query = Obliczenie.query
    
    if kod_filter:
        query = query.filter(Obliczenie.kod.like(f'%{kod_filter}%'))
    
    if data_od:
        try:
            data_od_obj = datetime.strptime(data_od, '%Y-%m-%d')
            query = query.filter(Obliczenie.data >= data_od_obj)
        except ValueError:
            pass
    
    if data_do:
        try:
            data_do_obj = datetime.strptime(data_do, '%Y-%m-%d')
            # Dodajemy jeden dzień, żeby data_do była włączona
            from datetime import timedelta
            data_do_obj = data_do_obj + timedelta(days=1)
            query = query.filter(Obliczenie.data < data_do_obj)
        except ValueError:
            pass
    
    obliczenia = query.order_by(Obliczenie.data.desc()).all()
    
    return render_template('calculation/historia.html', obliczenia=obliczenia, 
                           kod_filter=kod_filter, data_od=data_od, data_do=data_do)

@bp.route('/eksport/<int:obliczenie_id>')
def eksport(obliczenie_id):
    obliczenie = Obliczenie.query.get_or_404(obliczenie_id)
    grupa = obliczenie.grupa
    # Sortujemy metody alfabetycznie dla stałej kolejności
    metody_w_grupie = sorted(grupa.metody, key=lambda m: m.nazwa)
    wyniki = {w['nazwa']: w for w in obliczenie.wyniki()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Obliczenie"

    naglowki = ['Kod produktu', 'Grupa', 'Przedział', 'Data', 'Czas całkowity (min)']
    for metoda in metody_w_grupie:
        nazwa = metoda.nazwa
        naglowki.extend([f"{nazwa} - metry (mtr)", f"{nazwa} - pracownicy", f"{nazwa} - czas (min)"])
    ws.append(naglowki)

    for col in range(1, len(naglowki)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    wiersz = [
        obliczenie.kod,
        obliczenie.grupa.nazwa,
        obliczenie.przedzial,
        obliczenie.data.strftime('%Y-%m-%d %H:%M'),
        obliczenie.czas_calkowity
    ]
    for metoda in metody_w_grupie:
        if metoda.nazwa in wyniki:
            w = wyniki[metoda.nazwa]
            wiersz.extend([w['metry'], w['pracownicy'], w['czas_calkowity']])
        else:
            wiersz.extend([0, 0, 0])
    ws.append(wiersz)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'obliczenie_{obliczenie.id}_{obliczenie.kod}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/eksport_wszystkie')
def eksport_wszystkie():
    # Pobieramy wszystkie unikalne nazwy metod z bazy
    wszystkie_metody_db = db.session.query(Metoda.nazwa).distinct().all()
    WSZYSTKIE_METODY = [m[0] for m in wszystkie_metody_db]
    # Fallback na domyślne, gdyby baza była pusta
    if not WSZYSTKIE_METODY:
        WSZYSTKIE_METODY = list(Metoda.domyslne_czasy().keys())
    # Sortujemy alfabetycznie
    WSZYSTKIE_METODY.sort()

    obliczenia = Obliczenie.query.order_by(Obliczenie.data.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wszystkie obliczenia"

    naglowki = ['ID', 'Data', 'Kod', 'Grupa', 'Przedział', 'Czas całkowity (min)']
    for metoda_nazwa in WSZYSTKIE_METODY:
        naglowki.extend([f"{metoda_nazwa} - metry (mtr)", f"{metoda_nazwa} - pracownicy", f"{metoda_nazwa} - czas (min)"])
    ws.append(naglowki)

    for col in range(1, len(naglowki)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)

    for o in obliczenia:
        wyniki = {w['nazwa']: w for w in o.wyniki()}
        wiersz = [
            o.id,
            o.data.strftime('%Y-%m-%d %H:%M'),
            o.kod,
            o.grupa.nazwa,
            o.przedzial,
            o.czas_calkowity
        ]
        for metoda_nazwa in WSZYSTKIE_METODY:
            if metoda_nazwa in wyniki:
                w = wyniki[metoda_nazwa]
                wiersz.extend([w['metry'], w['pracownicy'], w['czas_calkowity']])
            else:
                wiersz.extend([0, 0, 0])
        ws.append(wiersz)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='wszystkie_obliczenia.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/usun/<int:obliczenie_id>', methods=['POST'])
def usun(obliczenie_id):
    try:
        obliczenie = Obliczenie.query.get_or_404(obliczenie_id)
        db.session.delete(obliczenie)
        db.session.commit()
        flash('Obliczenie zostało usunięte.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Błąd podczas usuwania: {str(e)}', 'danger')
    return redirect(url_for('calculation.historia'))